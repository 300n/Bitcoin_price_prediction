"""
Bitcoin Price Events Scraper
Collects major historical events that influenced Bitcoin's price
and stores them in an Excel file for price prediction analysis.
"""

import time
import re
import logging
from datetime import datetime
from dataclasses import dataclass, field

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

OUTPUT_FILE = "bitcoin_events.xlsx"


@dataclass
class BTCEvent:
    date: str
    title: str
    description: str
    category: str
    impact: str          # BULLISH / BEARISH / NEUTRAL
    source: str
    price_before: str = ""
    price_after: str = ""
    price_change_pct: str = ""
    url: str = ""
    tags: str = ""


# ---------------------------------------------------------------------------
# CURATED HISTORICAL DATASET
# A high-quality seed of every major documented event + approximate prices.
# Sources: CoinDesk, Investopedia, Wikipedia, Bitcoin Magazine archives.
# ---------------------------------------------------------------------------

HISTORICAL_EVENTS: list[BTCEvent] = [
    # 2009
    BTCEvent("2009-01-03", "Bitcoin Genesis Block Mined",
             "Satoshi Nakamoto mines the first Bitcoin block (Block 0), embedding "
             "'The Times 03/Jan/2009 Chancellor on brink of second bailout for banks'.",
             "Technology", "BULLISH", "Bitcoin Whitepaper / Genesis Block",
             "0", "0", "0", "", "genesis,satoshi,creation"),
    BTCEvent("2009-01-12", "First Bitcoin Transaction",
             "Satoshi sends 10 BTC to Hal Finney in the first ever peer-to-peer "
             "Bitcoin transaction.",
             "Adoption", "BULLISH", "Bitcoin History",
             "~0", "~0", "0", "", "first transaction,hal finney"),

    # 2010
    BTCEvent("2010-05-22", "Bitcoin Pizza Day",
             "Laszlo Hanyecz pays 10,000 BTC for two pizzas — the first real-world "
             "commercial transaction. Establishes a price of ~$0.0025/BTC.",
             "Adoption", "BULLISH", "Bitcoin Talk Forum",
             "~0.0025", "~0.0025", "0", "", "pizza,first purchase,adoption"),
    BTCEvent("2010-07-17", "Mt. Gox Exchange Launches",
             "Mt. Gox opens as the first major Bitcoin exchange, dramatically increasing "
             "liquidity and accessibility. Price jumps from $0.05 to $0.08 in days.",
             "Exchange", "BULLISH", "Mt. Gox / CoinDesk",
             "0.05", "0.08", "+60", "", "exchange,mt gox,liquidity"),
    BTCEvent("2010-08-15", "Bitcoin Value Overflow Incident",
             "A bug creates 184 billion BTC out of thin air. The transaction is reversed "
             "within hours via a hard fork. First serious protocol bug.",
             "Security", "BEARISH", "Bitcoin Talk / CVE",
             "0.07", "0.05", "-28", "", "bug,overflow,hard fork"),

    # 2011
    BTCEvent("2011-02-09", "Bitcoin Reaches $1",
             "Bitcoin hits $1 for the first time on Mt. Gox, a milestone widely covered "
             "by tech media.",
             "Price Milestone", "BULLISH", "Mt. Gox / TIME Magazine",
             "0.30", "1.00", "+233", "", "milestone,$1"),
    BTCEvent("2011-04-16", "TIME Magazine Article",
             "TIME Magazine publishes 'Online Cash Bitcoin Could Challenge Governments, "
             "Banks', massively increasing public awareness.",
             "Media", "BULLISH", "TIME Magazine",
             "0.90", "8.00", "+789", "", "media,mainstream,time"),
    BTCEvent("2011-06-19", "Mt. Gox Hack ($8.75)",
             "Mt. Gox is hacked; attacker manipulates the order book, crashing price from "
             "$17.50 to $0.01 momentarily. 650,000 BTC stolen over time.",
             "Security / Hack", "BEARISH", "Mt. Gox / Bitcoin Talk",
             "17.50", "7.00", "-60", "", "hack,mt gox,security"),
    BTCEvent("2011-11-20", "Silk Road Attention",
             "Senators Schumer and Manchin call on the DEA to shut down Silk Road, the "
             "first high-profile regulatory threat using Bitcoin.",
             "Regulation", "BEARISH", "US Senate / CNN",
             "3.00", "2.20", "-26", "", "silk road,regulation,darknet"),

    # 2012
    BTCEvent("2012-11-28", "First Bitcoin Halving",
             "Block reward halves from 50 BTC to 25 BTC. Historically the start of a "
             "long-term bull cycle.",
             "Halving", "BULLISH", "Bitcoin Protocol",
             "12.00", "13.00", "+8",  "", "halving,supply,mining"),

    # 2013
    BTCEvent("2013-03-18", "Cyprus Bailout / Banking Crisis",
             "EU imposes a 10% levy on Cyprus bank deposits. Bitcoin seen as a safe haven "
             "and surges from $45 to $266 in weeks.",
             "Macro / Banking", "BULLISH", "Reuters / CoinDesk",
             "45", "266", "+491", "", "banking crisis,safe haven,cyprus"),
    BTCEvent("2013-04-10", "First Major Crash from $266",
             "After a parabolic rise Bitcoin crashes 83% in 2 days due to Mt. Gox "
             "overload and profit taking.",
             "Market", "BEARISH", "Mt. Gox / CoinDesk",
             "266", "54", "-80", "", "crash,bubble,mt gox"),
    BTCEvent("2013-10-02", "Silk Road Shut Down by FBI",
             "FBI seizes Silk Road and arrests Ross Ulbricht. 26,000 BTC confiscated. "
             "Short-term bearish but long-term neutral as it removes criminal stigma.",
             "Regulation / Law", "BEARISH", "FBI / Reuters",
             "140", "109", "-22", "", "silk road,fbi,regulation"),
    BTCEvent("2013-11-29", "Bitcoin Reaches $1,000",
             "Bitcoin hits $1,000 for the first time. China accounts for large share of "
             "trading; widespread mainstream media coverage.",
             "Price Milestone", "BULLISH", "CoinDesk / CNN",
             "200", "1242", "+521", "", "milestone,$1000,all time high"),
    BTCEvent("2013-12-05", "China PBOC Bans Banks from Bitcoin",
             "People's Bank of China prohibits financial institutions from handling Bitcoin "
             "transactions, triggering a major sell-off.",
             "Regulation", "BEARISH", "PBOC / Bloomberg",
             "1100", "600", "-45", "", "china,regulation,ban"),

    # 2014
    BTCEvent("2014-02-07", "Mt. Gox Halts Withdrawals",
             "Mt. Gox suspends all withdrawals citing 'transaction malleability'. Marks "
             "start of its collapse.",
             "Exchange / Hack", "BEARISH", "Mt. Gox",
             "800", "550", "-31", "", "mt gox,withdrawal,collapse"),
    BTCEvent("2014-02-24", "Mt. Gox Files Bankruptcy",
             "Mt. Gox files for bankruptcy protection after losing ~850,000 BTC (~$450M). "
             "Largest Bitcoin exchange collapse.",
             "Exchange / Hack", "BEARISH", "Mt. Gox / WSJ",
             "550", "420", "-23", "", "mt gox,bankruptcy,hack"),
    BTCEvent("2014-10-06", "Bearish Trend Deepens",
             "Bitcoin enters a prolonged bear market, falling below $300 as Mt. Gox "
             "fallout and regulatory uncertainty persist.",
             "Market", "BEARISH", "CoinDesk",
             "420", "290", "-31", "", "bear market,mt gox,regulation"),

    # 2015
    BTCEvent("2015-01-14", "Coinbase Raises $75M Series C",
             "Coinbase raises $75M from NYSE, BBVA, and others — institutional money "
             "enters Bitcoin for the first time at scale.",
             "Institutional", "BULLISH", "Coinbase / Forbes",
             "175", "215", "+23", "", "coinbase,institutional,funding"),
    BTCEvent("2015-08-01", "Bitcoin XT Fork Proposal",
             "Bitcoin XT, a contentious hard fork proposal by Gavin Andresen and Mike "
             "Hearn to increase block size, creates community discord.",
             "Technology / Fork", "BEARISH", "Bitcoin Magazine",
             "280", "260", "-7", "", "fork,block size,scaling"),

    # 2016
    BTCEvent("2016-05-22", "Ethereum DAO Raises $150M",
             "The DAO raises $150M in Ether, diverting attention and capital from Bitcoin "
             "temporarily.",
             "Competition", "NEUTRAL", "CoinDesk / Ethereum",
             "450", "440", "-2", "", "ethereum,dao,competition"),
    BTCEvent("2016-07-09", "Second Bitcoin Halving",
             "Block reward halves from 25 BTC to 12.5 BTC. Bull cycle begins; price "
             "rises from ~$650 to $20,000 over 18 months.",
             "Halving", "BULLISH", "Bitcoin Protocol",
             "650", "660", "+2", "", "halving,supply,mining"),
    BTCEvent("2016-08-02", "Bitfinex Hack — 120,000 BTC Stolen",
             "Bitfinex exchange is hacked for 119,756 BTC (~$72M). Price drops 20% "
             "immediately.",
             "Security / Hack", "BEARISH", "Bitfinex / Reuters",
             "600", "480", "-20", "", "hack,bitfinex,security"),

    # 2017
    BTCEvent("2017-03-10", "SEC Rejects Winklevoss Bitcoin ETF",
             "SEC rejects the first Bitcoin ETF application by the Winklevoss twins, "
             "citing lack of market regulation.",
             "Regulation / ETF", "BEARISH", "SEC / Bloomberg",
             "1290", "1050", "-19", "", "etf,sec,regulation,winklevoss"),
    BTCEvent("2017-05-25", "Japan Recognizes Bitcoin as Legal Tender",
             "Japan's Payment Services Act officially recognizes Bitcoin as a legal payment "
             "method, triggering a major rally.",
             "Regulation / Adoption", "BULLISH", "Japanese FSA / Reuters",
             "2200", "2700", "+23", "", "japan,legal tender,adoption,regulation"),
    BTCEvent("2017-08-01", "Bitcoin Cash Hard Fork",
             "Bitcoin forks into Bitcoin (BTC) and Bitcoin Cash (BCH). BCH gives holders "
             "free coins; short-term volatility ensues.",
             "Technology / Fork", "NEUTRAL", "Bitcoin Magazine / CoinDesk",
             "2700", "2700", "0", "", "hard fork,bitcoin cash,bch,scaling"),
    BTCEvent("2017-09-04", "China Bans ICOs and Exchanges",
             "Chinese government bans ICOs and orders all domestic cryptocurrency exchanges "
             "to shut down. BTC drops ~40%.",
             "Regulation", "BEARISH", "PBOC / Bloomberg",
             "4900", "3000", "-39", "", "china,ban,ico,exchange,regulation"),
    BTCEvent("2017-10-31", "CME Announces Bitcoin Futures",
             "CME Group announces it will launch Bitcoin futures, legitimizing Bitcoin as "
             "a tradable financial asset.",
             "Institutional / Derivatives", "BULLISH", "CME / CNBC",
             "6100", "6400", "+5", "", "futures,cme,institutional,derivatives"),
    BTCEvent("2017-12-17", "Bitcoin ATH — $19,783",
             "Bitcoin reaches its then-all-time high of $19,783. Media frenzy; retail "
             "FOMO at peak. Start of 2018 bear market.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "15000", "19783", "+32", "", "all time high,ath,bubble,retail"),
    BTCEvent("2017-12-18", "CME Bitcoin Futures Launch",
             "CME launches cash-settled Bitcoin futures. Institutional short selling begins; "
             "marks the top of the 2017 bull run.",
             "Institutional / Derivatives", "BEARISH", "CME",
             "19500", "18000", "-8", "", "futures,cme,top,derivatives"),

    # 2018
    BTCEvent("2018-01-16", "Global Crackdown on Cryptocurrency",
             "South Korea, China, US regulators all signal intent to regulate. Market-wide "
             "crash begins; BTC drops from $14K to under $10K.",
             "Regulation", "BEARISH", "Reuters / Bloomberg",
             "14000", "9500", "-32", "", "regulation,crackdown,korea,global"),
    BTCEvent("2018-03-07", "SEC Issues Subpoenas to ICO Projects",
             "SEC issues subpoenas to dozens of ICO projects, signaling aggressive "
             "enforcement and spooking the market.",
             "Regulation", "BEARISH", "SEC / WSJ",
             "10500", "8900", "-15", "", "sec,ico,regulation,enforcement"),
    BTCEvent("2018-11-15", "Bitcoin Cash Hash War",
             "Craig Wright's Bitcoin SV vs. Roger Ver's Bitcoin ABC hash war. Miners "
             "divert hash power; BTC drops 50% in weeks.",
             "Technology / Fork", "BEARISH", "CoinDesk",
             "6300", "3500", "-44", "", "hash war,bitcoin sv,bitcoin cash,fork"),
    BTCEvent("2018-12-15", "Bitcoin Year Low — $3,122",
             "Bitcoin bottoms at $3,122, an 84% decline from the $19,783 ATH. Crypto "
             "winter sets in.",
             "Market", "BEARISH", "CoinDesk",
             "3500", "3122", "-11", "", "bear market,crypto winter,bottom"),

    # 2019
    BTCEvent("2019-06-18", "Facebook Announces Libra",
             "Facebook announces its Libra cryptocurrency project, bringing massive "
             "mainstream attention to crypto and Bitcoin.",
             "Institutional / Competition", "BULLISH", "Facebook / NYT",
             "8000", "11000", "+38", "", "facebook,libra,mainstream,adoption"),
    BTCEvent("2019-06-26", "Bitcoin Reaches $13,880 — 2019 High",
             "BTC rallies to $13,880, its highest since 2018 crash, driven by Libra "
             "announcement and renewed institutional interest.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "11000", "13880", "+26", "", "rally,2019 high,institutional"),

    # 2020
    BTCEvent("2020-03-12", "COVID-19 Black Thursday Crash",
             "Global markets crash due to COVID-19 pandemic. Bitcoin drops from $9,000 "
             "to $3,800 in 24 hours (−57%), one of its worst single-day drops.",
             "Macro / Pandemic", "BEARISH", "CoinDesk / Bloomberg",
             "9000", "3800", "-57", "", "covid,pandemic,crash,black thursday,macro"),
    BTCEvent("2020-05-11", "Third Bitcoin Halving",
             "Block reward halves from 12.5 to 6.25 BTC. Triggers the 2020-2021 bull run "
             "that takes BTC from $8,500 to $69,000.",
             "Halving", "BULLISH", "Bitcoin Protocol",
             "8500", "8800", "+4", "", "halving,supply,mining,bull run"),
    BTCEvent("2020-08-11", "MicroStrategy Buys $250M in Bitcoin",
             "MicroStrategy becomes the first public company to adopt Bitcoin as its "
             "primary treasury reserve asset.",
             "Institutional", "BULLISH", "MicroStrategy / Bloomberg",
             "11500", "12100", "+5", "", "microstrategy,institutional,corporate,treasury"),
    BTCEvent("2020-10-21", "PayPal Enables Bitcoin Buying",
             "PayPal announces all 346 million users can buy, sell, and hold Bitcoin "
             "directly. Massive mainstream adoption catalyst.",
             "Adoption", "BULLISH", "PayPal / CNBC",
             "12000", "13200", "+10", "", "paypal,mainstream,adoption,retail"),
    BTCEvent("2020-12-16", "Bitcoin Breaks 2017 ATH — $20,000",
             "Bitcoin breaks its 2017 all-time high of ~$20,000 for the first time, "
             "validating the 2020 bull cycle.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "18000", "20000", "+11", "", "all time high,ath,milestone,breakout"),

    # 2021
    BTCEvent("2021-01-29", "Elon Musk Adds #Bitcoin to Twitter Bio",
             "Elon Musk adds '#bitcoin' to his Twitter bio. BTC surges 20% in hours, "
             "demonstrating social media influence on price.",
             "Social Media / Influencer", "BULLISH", "Twitter / CoinDesk",
             "32000", "38000", "+19", "", "elon musk,twitter,social media,influencer"),
    BTCEvent("2021-02-08", "Tesla Buys $1.5B in Bitcoin",
             "Tesla announces it purchased $1.5B in Bitcoin and will accept it as payment. "
             "BTC surges to $44K.",
             "Institutional", "BULLISH", "Tesla / SEC Filing",
             "38000", "44000", "+16", "", "tesla,institutional,corporate,elon musk"),
    BTCEvent("2021-02-19", "Bitcoin Hits $52,640 — New ATH",
             "Bitcoin sets new all-time high driven by Tesla purchase, PayPal, "
             "institutional demand, and stimulus checks.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "47000", "52640", "+12", "", "all time high,ath,institutional,rally"),
    BTCEvent("2021-04-14", "Coinbase Direct Listing on NASDAQ",
             "Coinbase goes public on NASDAQ at $328/share (COIN). Bitcoin hits $64,000 "
             "same day — landmark for crypto legitimacy.",
             "Institutional / Exchange", "BULLISH", "NASDAQ / Bloomberg",
             "60000", "64895", "+8", "", "coinbase,ipo,nasdaq,institutional,adoption"),
    BTCEvent("2021-05-12", "Tesla Stops Accepting Bitcoin — Environmental",
             "Elon Musk announces Tesla will stop accepting Bitcoin due to environmental "
             "concerns over mining. BTC drops 15% immediately.",
             "Adoption / Environmental", "BEARISH", "Elon Musk / Twitter",
             "54000", "46000", "-15", "", "tesla,elon musk,environment,mining,esg"),
    BTCEvent("2021-05-19", "China Bans Bitcoin Mining",
             "China cracks down on Bitcoin mining and trading. Combined with Tesla news, "
             "BTC crashes from $64K to $30K (−53%).",
             "Regulation / Mining", "BEARISH", "PBOC / Bloomberg",
             "43000", "30000", "-30", "", "china,mining ban,regulation,crash"),
    BTCEvent("2021-07-21", "Bitcoin Bottoms at $29,700 then Recovers",
             "BTC hits cycle low of $29,700 then surges as hash rate recovers post-China "
             "ban and institutional buying resumes.",
             "Market", "BULLISH", "CoinDesk",
             "29700", "33000", "+11", "", "bottom,recovery,hash rate,institutional"),
    BTCEvent("2021-09-07", "El Salvador Makes Bitcoin Legal Tender",
             "El Salvador becomes the first country to adopt Bitcoin as legal tender. "
             "Chivo wallet launched. Short-term volatile but historically bullish.",
             "Regulation / Adoption", "BULLISH", "El Salvador Government / Reuters",
             "52000", "53000", "+2", "", "el salvador,legal tender,adoption,country"),
    BTCEvent("2021-10-20", "First Bitcoin Futures ETF (BITO) Launches",
             "ProShares Bitcoin Futures ETF (BITO) launches on NYSE — the first US-listed "
             "Bitcoin ETF. BTC hits new ATH of $67K.",
             "Institutional / ETF", "BULLISH", "ProShares / NYSE",
             "62000", "67000", "+8", "", "etf,bito,proshares,institutional,nyse"),
    BTCEvent("2021-11-10", "Bitcoin ATH — $69,000",
             "Bitcoin reaches its current all-time high of $68,789, driven by ETF launch, "
             "institutional demand, and inflation fears.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "67000", "68789", "+3", "", "all time high,ath,$69k,inflation"),

    # 2022
    BTCEvent("2022-01-21", "Crypto Market Sell-Off on Fed Rate Hike Fears",
             "Federal Reserve signals aggressive rate hike cycle. Risk assets sell off; "
             "BTC drops from $47K to $33K.",
             "Macro / Fed", "BEARISH", "Fed / Bloomberg",
             "47000", "33000", "-30", "", "fed,rate hike,macro,inflation,risk off"),
    BTCEvent("2022-05-09", "Luna/UST Collapse",
             "Terra Luna ecosystem collapses; UST de-pegs from $1 and Luna goes to zero. "
             "$60B wiped out; BTC drops from $36K to $26K.",
             "DeFi / Stablecoin", "BEARISH", "CoinDesk / Bloomberg",
             "36000", "26000", "-28", "", "luna,terra,ust,stablecoin,depeg,collapse"),
    BTCEvent("2022-06-13", "Celsius Pauses Withdrawals",
             "Celsius Network halts all withdrawals and transfers citing extreme market "
             "conditions. BTC drops below $20K.",
             "Exchange / Lending", "BEARISH", "Celsius / Bloomberg",
             "28000", "20000", "-29", "", "celsius,withdrawal freeze,lending,contagion"),
    BTCEvent("2022-06-18", "Three Arrows Capital Collapse",
             "Hedge fund Three Arrows Capital (3AC) defaults on loans worth $670M. "
             "Crypto credit crisis deepens.",
             "Institutional / Contagion", "BEARISH", "WSJ / Bloomberg",
             "20000", "17500", "-13", "", "three arrows,3ac,hedge fund,contagion,liquidation"),
    BTCEvent("2022-07-13", "Bitcoin Bottom — $17,593",
             "Bitcoin hits $17,593, the lowest since December 2020. Miners capitulate; "
             "hash ribbons signal extreme fear.",
             "Market", "BEARISH", "CoinDesk",
             "19000", "17593", "-7", "", "bottom,bear market,miners,capitulation"),
    BTCEvent("2022-11-08", "FTX Liquidity Crisis",
             "Binance CEO CZ tweets concern over FTT tokens; FTX faces bank run. Binance "
             "backs out of rescue deal. BTC drops from $21K to $16K.",
             "Exchange / Fraud", "BEARISH", "CoinDesk / WSJ",
             "21000", "16000", "-24", "", "ftx,sam bankman-fried,sbf,exchange,fraud,contagion"),
    BTCEvent("2022-11-11", "FTX Files Bankruptcy — Sam Bankman-Fried Arrested",
             "FTX files for Chapter 11 bankruptcy. SBF arrested. $8B in customer funds "
             "missing. Largest crypto exchange collapse. BTC hits $15,600.",
             "Exchange / Fraud", "BEARISH", "FTX / DOJ",
             "16000", "15600", "-3", "", "ftx,sbf,bankruptcy,fraud,contagion"),

    # 2023
    BTCEvent("2023-01-12", "Bitcoin Starts 2023 Recovery",
             "Bitcoin rallies from $16K lows, starting the 2023 recovery rally amid "
             "expectations of Fed pivot and post-FTX reset.",
             "Market", "BULLISH", "CoinDesk",
             "17000", "21000", "+24", "", "recovery,2023,bear market bottom,fed pivot"),
    BTCEvent("2023-03-10", "Silicon Valley Bank Collapse",
             "SVB collapses in the second-largest US bank failure ever. Bitcoin initially "
             "drops then surges as banking fears resurface.",
             "Macro / Banking", "BULLISH", "FDIC / Bloomberg",
             "20000", "22000", "+10", "", "svb,banking crisis,safe haven,macro"),
    BTCEvent("2023-06-15", "BlackRock Files for Bitcoin Spot ETF",
             "BlackRock, the world's largest asset manager ($10T AUM), files for a spot "
             "Bitcoin ETF. Considered the most credible application to date.",
             "Institutional / ETF", "BULLISH", "BlackRock / SEC",
             "25000", "30000", "+20", "", "blackrock,etf,spot etf,institutional"),
    BTCEvent("2023-10-23", "Bitcoin Surges on ETF Approval Rumors",
             "Bitcoin surges to $35K on fake news of BlackRock ETF approval, then "
             "partially retraces but holds gains on real anticipation.",
             "ETF / Media", "BULLISH", "CoinDesk / Bloomberg",
             "27000", "35000", "+30", "", "etf,blackrock,approval,rumor"),

    # 2024
    BTCEvent("2024-01-10", "SEC Approves 11 Spot Bitcoin ETFs",
             "SEC simultaneously approves 11 spot Bitcoin ETFs including BlackRock IBIT, "
             "Fidelity FBTC. First day sees $4.6B in trading volume.",
             "Institutional / ETF", "BULLISH", "SEC / Bloomberg",
             "46000", "47000", "+2", "", "spot etf,sec,blackrock,fidelity,ibit,institutional"),
    BTCEvent("2024-02-14", "Bitcoin ETF Inflows Surge — IBIT Record",
             "BlackRock's IBIT ETF records $612M in a single day, the largest daily "
             "inflow of any ETF ever at the time.",
             "Institutional / ETF", "BULLISH", "BlackRock / Bloomberg",
             "52000", "54000", "+4", "", "etf,blackrock,ibit,inflows,institutional"),
    BTCEvent("2024-03-05", "Bitcoin Breaks 2021 ATH — $69,000+",
             "Bitcoin surpasses its 2021 all-time high of $69K for the first time, "
             "driven by ETF inflows and halving anticipation.",
             "Price Milestone", "BULLISH", "CoinDesk",
             "62000", "69200", "+12", "", "all time high,ath,breakout,etf,halving"),
    BTCEvent("2024-04-19", "Fourth Bitcoin Halving",
             "Block reward halves from 6.25 to 3.125 BTC at block 840,000. Miner fees "
             "spike to record highs due to Runes protocol launch.",
             "Halving", "BULLISH", "Bitcoin Protocol",
             "64000", "64000", "0", "", "halving,supply,mining,runes,ordinals"),
    BTCEvent("2024-05-23", "SEC Approves Spot Ethereum ETFs",
             "SEC unexpectedly approves spot Ethereum ETFs, signaling regulatory shift "
             "and benefiting the broader crypto market including BTC.",
             "Institutional / ETF / Regulation", "BULLISH", "SEC / Bloomberg",
             "67000", "69000", "+3", "", "ethereum,etf,sec,regulation,approval"),
    BTCEvent("2024-10-14", "Bitcoin Surges to $68K — Election Optimism",
             "Bitcoin rises to $68K driven by expectations that a Trump presidency would "
             "be pro-crypto.",
             "Political / Macro", "BULLISH", "CoinDesk",
             "62000", "68000", "+10", "", "trump,election,political,regulation,pro-crypto"),
    BTCEvent("2024-11-06", "Trump Wins US Presidential Election",
             "Donald Trump wins the 2024 US election. Bitcoin surges past $75K as markets "
             "price in crypto-friendly policies and potential strategic reserve.",
             "Political / Regulation", "BULLISH", "AP News / CoinDesk",
             "68000", "75000", "+10", "", "trump,election,president,strategic reserve,pro-crypto"),
    BTCEvent("2024-11-22", "Bitcoin Breaks $100,000 for First Time",
             "Bitcoin surpasses $100,000 for the first time, driven by Trump victory, "
             "ETF inflows, and halving supply squeeze.",
             "Price Milestone", "BULLISH", "CoinDesk / Bloomberg",
             "90000", "103000", "+14", "", "milestone,$100k,all time high,ath"),

    # 2025
    BTCEvent("2025-01-20", "Trump Signs Crypto Executive Order",
             "President Trump signs an executive order on Day 1 directing the creation of "
             "a Presidential Digital Asset Working Group and potential strategic BTC reserve.",
             "Regulation / Political", "BULLISH", "White House / CoinDesk",
             "102000", "107000", "+5", "", "trump,executive order,strategic reserve,regulation"),
    BTCEvent("2025-03-07", "US Strategic Bitcoin Reserve Announced",
             "Trump signs executive order establishing a US Strategic Bitcoin Reserve "
             "using seized government BTC (~200,000 BTC). BTC surges on the news.",
             "Regulation / Government", "BULLISH", "White House / Bloomberg",
             "88000", "95000", "+8", "", "strategic reserve,government,us,trump,adoption"),
]


# ---------------------------------------------------------------------------
# LIVE SCRAPERS (supplemental — adds recent news)
# ---------------------------------------------------------------------------

def scrape_coindesk_news(max_articles: int = 20) -> list[BTCEvent]:
    """Scrape recent Bitcoin headlines from CoinDesk."""
    events: list[BTCEvent] = []
    url = "https://www.coindesk.com/tag/bitcoin"
    try:
        log.info("Scraping CoinDesk...")
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "lxml")

        articles = soup.select("a[data-testid='card-title-link'], "
                               "h4 a, h3 a, .article-card a, "
                               "[class*='headline'] a")[:max_articles]
        seen: set[str] = set()
        for a in articles:
            title = a.get_text(strip=True)
            href = a.get("href", "")
            if not title or title in seen or len(title) < 20:
                continue
            seen.add(title)
            full_url = href if href.startswith("http") else f"https://www.coindesk.com{href}"
            impact = classify_impact(title)
            events.append(BTCEvent(
                date=datetime.today().strftime("%Y-%m-%d"),
                title=title[:200],
                description="[Scraped from CoinDesk — description requires full article fetch]",
                category=infer_category(title),
                impact=impact,
                source="CoinDesk",
                url=full_url,
                tags=extract_tags(title),
            ))
        log.info(f"CoinDesk: {len(events)} articles found")
    except Exception as e:
        log.warning(f"CoinDesk scrape failed: {e}")
    return events


def scrape_bitcoin_magazine(max_articles: int = 20) -> list[BTCEvent]:
    """Scrape recent news from Bitcoin Magazine."""
    events: list[BTCEvent] = []
    url = "https://bitcoinmagazine.com/articles"
    try:
        log.info("Scraping Bitcoin Magazine...")
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "lxml")

        articles = soup.select("h2 a, h3 a, .article__title a, "
                               "[class*='title'] a")[:max_articles]
        seen: set[str] = set()
        for a in articles:
            title = a.get_text(strip=True)
            href = a.get("href", "")
            if not title or title in seen or len(title) < 20:
                continue
            seen.add(title)
            full_url = href if href.startswith("http") else f"https://bitcoinmagazine.com{href}"
            events.append(BTCEvent(
                date=datetime.today().strftime("%Y-%m-%d"),
                title=title[:200],
                description="[Scraped from Bitcoin Magazine]",
                category=infer_category(title),
                impact=classify_impact(title),
                source="Bitcoin Magazine",
                url=full_url,
                tags=extract_tags(title),
            ))
        log.info(f"Bitcoin Magazine: {len(events)} articles found")
    except Exception as e:
        log.warning(f"Bitcoin Magazine scrape failed: {e}")
    return events


def scrape_cointelegraph(max_articles: int = 20) -> list[BTCEvent]:
    """Scrape recent Bitcoin news from CoinTelegraph."""
    events: list[BTCEvent] = []
    url = "https://cointelegraph.com/tags/bitcoin"
    try:
        log.info("Scraping CoinTelegraph...")
        r = requests.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        soup = BeautifulSoup(r.text, "lxml")

        articles = soup.select("a.post-card__title-link, h2 a, h3 a, "
                               "[class*='post-card'] a")[:max_articles]
        seen: set[str] = set()
        for a in articles:
            title = a.get_text(strip=True)
            href = a.get("href", "")
            if not title or title in seen or len(title) < 20:
                continue
            seen.add(title)
            full_url = href if href.startswith("http") else f"https://cointelegraph.com{href}"
            events.append(BTCEvent(
                date=datetime.today().strftime("%Y-%m-%d"),
                title=title[:200],
                description="[Scraped from CoinTelegraph]",
                category=infer_category(title),
                impact=classify_impact(title),
                source="CoinTelegraph",
                url=full_url,
                tags=extract_tags(title),
            ))
        log.info(f"CoinTelegraph: {len(events)} articles found")
    except Exception as e:
        log.warning(f"CoinTelegraph scrape failed: {e}")
    return events


# ---------------------------------------------------------------------------
# NLP HELPERS
# ---------------------------------------------------------------------------

BULLISH_KEYWORDS = {
    "buy", "bull", "surge", "rally", "soar", "jump", "rise", "gain", "record",
    "ath", "all-time high", "approve", "approval", "adopt", "adoption", "launch",
    "invest", "institutional", "etf", "halving", "legal", "green", "recover",
    "breakout", "milestone", "fund", "reserve", "trump", "pro-crypto", "listing",
    "upgrade", "partnership", "integration", "accept",
}

BEARISH_KEYWORDS = {
    "sell", "bear", "crash", "drop", "fall", "plunge", "dump", "ban", "hack",
    "stolen", "theft", "fraud", "scam", "bankrupt", "suspend", "halt", "block",
    "crackdown", "restrict", "fine", "arrest", "collapse", "lose", "loss",
    "fear", "panic", "FUD", "warning", "risk", "regulation", "sue", "lawsuit",
    "reject", "denied", "debt", "contagion", "liquidat",
}


def classify_impact(text: str) -> str:
    text_lower = text.lower()
    bull = sum(1 for k in BULLISH_KEYWORDS if k in text_lower)
    bear = sum(1 for k in BEARISH_KEYWORDS if k in text_lower)
    if bull > bear:
        return "BULLISH"
    if bear > bull:
        return "BEARISH"
    return "NEUTRAL"


CATEGORY_MAP = {
    "hack|theft|stolen|breach|exploit": "Security / Hack",
    "etf|fund|futures|derivative|option": "Institutional / ETF",
    "halving|block reward|mining|hash": "Halving / Mining",
    "regulation|ban|law|legal|sec|cftc|fca|pboc|government": "Regulation",
    "exchange|coinbase|binance|kraken|ftx|bybit|okx": "Exchange",
    "macro|fed|inflation|rate|recession|bank|crisis": "Macro",
    "fork|upgrade|protocol|segwit|taproot|lightning": "Technology / Fork",
    "institutional|corporate|microstrategy|tesla|blackrock": "Institutional",
    "adoption|payment|accept|merchant": "Adoption",
    "defi|stablecoin|usdc|usdt|dai|ust": "DeFi / Stablecoin",
    "social|twitter|tweet|elon|influencer": "Social Media / Influencer",
    "political|president|election|congress|senate": "Political",
}


def infer_category(text: str) -> str:
    text_lower = text.lower()
    for pattern, category in CATEGORY_MAP.items():
        if re.search(pattern, text_lower):
            return category
    return "Market"


TAG_KEYWORDS = [
    "bitcoin", "btc", "etf", "sec", "halving", "institutional", "regulation",
    "exchange", "mining", "defi", "stablecoin", "macro", "fed", "trump",
    "blackrock", "coinbase", "tesla", "china", "fork", "lightning", "taproot",
    "crash", "rally", "adoption", "ban", "hack", "fraud",
]


def extract_tags(text: str) -> str:
    text_lower = text.lower()
    return ",".join(t for t in TAG_KEYWORDS if t in text_lower)


# ---------------------------------------------------------------------------
# EXCEL BUILDER
# ---------------------------------------------------------------------------

CATEGORY_COLORS = {
    "Halving": "FFF3CD",
    "Halving / Mining": "FFF3CD",
    "Price Milestone": "D4EDDA",
    "Institutional": "CCE5FF",
    "Institutional / ETF": "CCE5FF",
    "Institutional / Derivatives": "CCE5FF",
    "Institutional / Exchange": "CCE5FF",
    "Institutional / Contagion": "F8D7DA",
    "Regulation": "E2E3E5",
    "Regulation / Adoption": "E2E3E5",
    "Regulation / Law": "E2E3E5",
    "Regulation / ETF": "E2E3E5",
    "Regulation / Mining": "E2E3E5",
    "Regulation / Government": "E2E3E5",
    "Regulation / Political": "E2E3E5",
    "Security / Hack": "F8D7DA",
    "Exchange / Hack": "F8D7DA",
    "Exchange / Lending": "F8D7DA",
    "Exchange / Fraud": "F8D7DA",
    "Technology / Fork": "E8D5F5",
    "Macro": "FFE5D0",
    "Macro / Banking": "FFE5D0",
    "Macro / Pandemic": "FFE5D0",
    "Macro / Fed": "FFE5D0",
    "DeFi / Stablecoin": "FCE4EC",
    "Social Media / Influencer": "E0F7FA",
    "Adoption": "F0FFF0",
    "Market": "F5F5F5",
    "Political": "FFF8E1",
    "Political / Regulation": "FFF8E1",
    "Political / Macro": "FFF8E1",
    "Competition": "FAFAFA",
    "Media": "E3F2FD",
}

IMPACT_COLORS = {
    "BULLISH": "C6EFCE",
    "BEARISH": "FFC7CE",
    "NEUTRAL": "FFEB9C",
}

THIN = Side(border_style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("Date", 14),
    ("Title", 55),
    ("Category", 28),
    ("Impact", 10),
    ("Price Before ($)", 17),
    ("Price After ($)", 16),
    ("Change (%)", 12),
    ("Description", 80),
    ("Source", 25),
    ("URL", 50),
    ("Tags", 40),
]


def _cell_fill(color_hex: str) -> PatternFill:
    return PatternFill("solid", fgColor=color_hex)


def build_excel(events: list[BTCEvent], output_path: str) -> None:
    log.info(f"Building Excel file: {output_path}")
    df = _events_to_df(events)

    # ---- Main events sheet ----
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="BTC Events", index=False)
        _style_events_sheet(writer.book["BTC Events"], df)

        # ---- Statistics sheet ----
        stats_df = _build_stats(df)
        stats_df.to_excel(writer, sheet_name="Statistics", index=False)
        _style_stats_sheet(writer.book["Statistics"])

        # ---- Category summary ----
        cat_df = _build_category_summary(df)
        cat_df.to_excel(writer, sheet_name="Category Summary", index=False)

        # ---- Prediction features ----
        feat_df = _build_prediction_features(df)
        feat_df.to_excel(writer, sheet_name="Prediction Features", index=False)
        _style_prediction_sheet(writer.book["Prediction Features"])

    log.info(f"Excel saved: {output_path}  ({len(events)} events)")


def _events_to_df(events: list[BTCEvent]) -> pd.DataFrame:
    rows = []
    for e in events:
        rows.append({
            "Date": e.date,
            "Title": e.title,
            "Category": e.category,
            "Impact": e.impact,
            "Price Before ($)": e.price_before,
            "Price After ($)": e.price_after,
            "Change (%)": e.price_change_pct,
            "Description": e.description,
            "Source": e.source,
            "URL": e.url,
            "Tags": e.tags,
        })
    df = pd.DataFrame(rows)
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.strftime("%Y-%m-%d")
    df = df.sort_values("Date").reset_index(drop=True)
    return df


def _style_events_sheet(ws, df: pd.DataFrame) -> None:
    # Header row
    header_fill = _cell_fill("1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # Data rows
    col_names = [c[0] for c in COLUMNS]
    for row_idx, row in df.iterrows():
        excel_row = row_idx + 2
        category = str(row.get("Category", ""))
        impact = str(row.get("Impact", ""))
        cat_color = CATEGORY_COLORS.get(category, "FFFFFF")
        impact_color = IMPACT_COLORS.get(impact, "FFFFFF")

        for col_idx, col_name in enumerate(col_names, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.value = row.get(col_name, "")
            cell.border = BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(col_name in ("Description", "Title", "Tags")),
            )
            if col_name == "Impact":
                cell.fill = _cell_fill(impact_color)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_name == "Category":
                cell.fill = _cell_fill(cat_color)
            elif col_name == "Change (%)":
                val = str(cell.value).replace("%", "").strip()
                try:
                    num = float(val)
                    cell.fill = _cell_fill("C6EFCE" if num >= 0 else "FFC7CE")
                    cell.font = Font(bold=True)
                except ValueError:
                    pass
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_name == "URL" and str(cell.value).startswith("http"):
                cell.hyperlink = str(cell.value)
                cell.font = Font(color="0563C1", underline="single")
            else:
                cell.fill = _cell_fill(cat_color)

        ws.row_dimensions[excel_row].height = 40

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions


def _build_stats(df: pd.DataFrame) -> pd.DataFrame:
    numeric_change = pd.to_numeric(
        df["Change (%)"].astype(str).str.replace("%", "").str.replace("+", ""),
        errors="coerce"
    )
    df = df.copy()
    df["_change"] = numeric_change

    stats = {
        "Metric": [
            "Total Events",
            "Bullish Events",
            "Bearish Events",
            "Neutral Events",
            "Avg Bullish Move (%)",
            "Avg Bearish Move (%)",
            "Largest Single Rally (%)",
            "Largest Single Drop (%)",
            "Most Common Category",
            "Events per Year (avg)",
            "Date Range Start",
            "Date Range End",
        ],
        "Value": [
            len(df),
            len(df[df["Impact"] == "BULLISH"]),
            len(df[df["Impact"] == "BEARISH"]),
            len(df[df["Impact"] == "NEUTRAL"]),
            round(df[df["_change"] > 0]["_change"].mean(), 1) if not df[df["_change"] > 0].empty else 0,
            round(df[df["_change"] < 0]["_change"].mean(), 1) if not df[df["_change"] < 0].empty else 0,
            round(df["_change"].max(), 1),
            round(df["_change"].min(), 1),
            df["Category"].value_counts().idxmax() if not df.empty else "N/A",
            round(len(df) / max(1, len(df["Date"].str[:4].unique())), 1),
            df["Date"].min(),
            df["Date"].max(),
        ],
    }
    return pd.DataFrame(stats)


def _style_stats_sheet(ws) -> None:
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 30
    header_fill = _cell_fill("1F3864")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BORDER


def _build_category_summary(df: pd.DataFrame) -> pd.DataFrame:
    summary = df.groupby(["Category", "Impact"]).size().reset_index(name="Count")
    return summary.sort_values("Count", ascending=False)


def _build_prediction_features(df: pd.DataFrame) -> pd.DataFrame:
    """
    Outputs a machine-learning ready feature table.
    Each row = one event with numeric features useful for prediction models.
    """
    df = df.copy()
    df["change_num"] = pd.to_numeric(
        df["Change (%)"].astype(str).str.replace("%", "").str.replace("+", ""),
        errors="coerce"
    )
    df["is_halving"] = df["Category"].str.contains("Halving", na=False).astype(int)
    df["is_regulation"] = df["Category"].str.contains("Regulation", na=False).astype(int)
    df["is_hack"] = df["Category"].str.contains("Hack|Security|Fraud", na=False).astype(int)
    df["is_institutional"] = df["Category"].str.contains("Institutional", na=False).astype(int)
    df["is_macro"] = df["Category"].str.contains("Macro", na=False).astype(int)
    df["is_exchange"] = df["Category"].str.contains("Exchange", na=False).astype(int)
    df["is_etf"] = df["Tags"].str.contains("etf", na=False).astype(int)
    df["is_china"] = df["Tags"].str.contains("china", na=False).astype(int)
    df["is_us_gov"] = df["Tags"].str.contains("trump|sec|regulation|government", na=False).astype(int)
    df["impact_numeric"] = df["Impact"].map({"BULLISH": 1, "NEUTRAL": 0, "BEARISH": -1})

    features = df[[
        "Date", "Title", "Category", "Impact", "impact_numeric",
        "change_num",
        "is_halving", "is_regulation", "is_hack", "is_institutional",
        "is_macro", "is_exchange", "is_etf", "is_china", "is_us_gov",
    ]].rename(columns={"change_num": "price_change_pct"})
    return features


def _style_prediction_sheet(ws) -> None:
    header_fill = _cell_fill("1F3864")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 28
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main() -> None:
    log.info("=== Bitcoin Price Events Scraper ===")

    all_events: list[BTCEvent] = list(HISTORICAL_EVENTS)
    log.info(f"Loaded {len(all_events)} curated historical events")

    # Scrape live news sources
    for scraper in [scrape_coindesk_news, scrape_bitcoin_magazine, scrape_cointelegraph]:
        live_events = scraper()
        all_events.extend(live_events)
        time.sleep(2)  # polite delay between requests

    # Deduplicate by title
    seen_titles: set[str] = set()
    unique_events: list[BTCEvent] = []
    for e in all_events:
        key = e.title.lower().strip()[:80]
        if key not in seen_titles:
            seen_titles.add(key)
            unique_events.append(e)

    log.info(f"Total unique events after deduplication: {len(unique_events)}")

    output_path = f"/Users/noledge/Insider-Trading/Insider_Trading/{OUTPUT_FILE}"
    build_excel(unique_events, output_path)

    # Print summary
    bullish = sum(1 for e in unique_events if e.impact == "BULLISH")
    bearish = sum(1 for e in unique_events if e.impact == "BEARISH")
    neutral = sum(1 for e in unique_events if e.impact == "NEUTRAL")

    print(f"\n{'='*55}")
    print(f"  Bitcoin Events Database — Build Complete")
    print(f"{'='*55}")
    print(f"  Total events  : {len(unique_events)}")
    print(f"  Bullish       : {bullish}")
    print(f"  Bearish       : {bearish}")
    print(f"  Neutral       : {neutral}")
    print(f"  Output file   : {output_path}")
    print(f"  Sheets        : BTC Events | Statistics | Category Summary | Prediction Features")
    print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
